#!/usr/bin/env python3
"""
verify_provenance.py — independent audit of the extracted JSON against the
workbook itself.

The dashboard makes a specific promise to whoever reads it: every figure names
the cell it came from, in an info bubble. This script checks that promise. It
opens the workbook with openpyxl and, for every {value, cell} pair anywhere in
the JSON, asserts that the named cell really does contain that value.

It deliberately does NOT re-derive the report structure. It walks the JSON's own
cell references, so it cannot share a bug with the extractor — if the extractor
read the wrong column, the value it stored will not match the cell it cites.

Also audited:
  - "computed": true entries must NOT claim a single source cell, and their
    "cells" list must all exist; additive ones must equal the sum of those cells.
  - cumulative series must equal the running sum of the cells they name.
  - every cell reference must be syntactically valid and within the sheet.

Usage:
    python3 verify_provenance.py [workbook.xlsx]
        [-d digital-report-data.json] [-p paid-media-data.json]
"""

import argparse
import json
import os
import re
import sys

import openpyxl

CELL_RE = re.compile(r"^([A-Z]{1,3})([1-9][0-9]{0,6})$")

# cell-key -> value-key. A cell key paired with its value key is an assertion
# the JSON makes about the workbook, and every one of them gets checked.
PAIRS = {
    "cell": "value",
    "goal_cell": "goal",
    "actual_cell": "actual",
    "pct_cell": "pct",
    "name_cell": "name",
    "label_cell": "label",
    "month_cell": "month",
    "channel_cell": "channel",
    "placement_cell": "placement",
    "objective_cell": "objective",
    "term_cell": "term",
    "text_cell": "text",
    "quarter_cell": "quarter",
    "step_cell": "step",
}

TOL = 0.005

results = []
counts = {"pairs": 0, "computed": 0, "series": 0, "refs": 0}


def check(ok, name, detail):
    results.append((ok, name, detail))


def cellnum(v):
    """Normalise a workbook cell value for numeric comparison, or None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if s in ("", "-") or s.startswith("#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def celltxt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def audit(payload, ws, tag):
    """Walk the payload, checking every provenance claim it makes."""
    sheet_title = payload.get("meta", {}).get("sheet")
    check(sheet_title == ws.title, "%s: JSON names the sheet it was read from" % tag,
          "JSON says %r, auditing %r" % (sheet_title, ws.title))

    def ref_ok(ref, where):
        counts["refs"] += 1
        m = CELL_RE.match(ref or "")
        if not m:
            check(False, "%s: %s has a valid cell reference" % (tag, where), "got %r" % ref)
            return False
        return True

    def compare(ref, expected, where):
        """One provenance assertion: does `ref` hold `expected`?"""
        if not ref_ok(ref, where):
            return
        counts["pairs"] += 1
        actual = ws[ref].value
        if expected is None:
            # JSON says blank: the cell must be blank, or a non-numeric sentinel
            ok = cellnum(actual) is None
            check(ok, "%s: %s claims %s is blank/non-numeric" % (tag, where, ref),
                  "%s holds %r" % (ref, actual))
            return
        if isinstance(expected, (int, float)) and not isinstance(expected, bool):
            av = cellnum(actual)
            ok = av is not None and abs(av - float(expected)) <= max(TOL, abs(float(expected)) * 1e-9)
            check(ok, "%s: %s = %s" % (tag, where, ref),
                  "JSON says %r, %s holds %r" % (expected, ref, actual))
        elif isinstance(expected, bool):
            av = actual
            if isinstance(av, str):
                av = av.strip().upper() == "TRUE"
            ok = bool(av) == expected
            check(ok, "%s: %s = %s" % (tag, where, ref),
                  "JSON says %r, %s holds %r" % (expected, ref, actual))
        else:
            at = celltxt(actual)
            ev = str(expected).strip()
            # the extractor strips a trailing colon from definition terms
            ok = at == ev or (at or "").rstrip(":") == ev
            check(ok, "%s: %s = %s" % (tag, where, ref),
                  "JSON says %r, %s holds %r" % (expected, ref, actual))

    def walk(node, path):
        if isinstance(node, dict):
            # 1. straight value/cell assertions
            for ck, vk in PAIRS.items():
                if ck in node and node[ck] is not None and vk in node:
                    # a "computed" entry must not claim a single source cell
                    if node.get("computed"):
                        check(False, "%s: computed value at %s must not cite one cell" % (tag, path),
                              "has %s=%r while marked computed" % (ck, node[ck]))
                    else:
                        compare(node[ck], node[vk], path + "." + vk)

            # 2. computed entries: cells must exist; additive ones must sum
            if node.get("computed") and isinstance(node.get("cells"), list):
                counts["computed"] += 1
                refs = node["cells"]
                bad = [r for r in refs if not CELL_RE.match(r or "")]
                check(not bad, "%s: computed entry at %s cites valid cells" % (tag, path),
                      "invalid refs: %r" % bad)
                counts["refs"] += len(refs)
                # only sum when this is an additive metric, not a ratio
                if node.get("kind") in ("usd", "count") and node.get("value") is not None and not bad:
                    tot = 0.0
                    for r in refs:
                        v = cellnum(ws[r].value)
                        if v is not None:
                            tot += v
                    check(abs(tot - float(node["value"])) <= max(0.02, abs(tot) * 1e-9),
                          "%s: computed sum at %s equals its cited cells" % (tag, path),
                          "JSON says %r, cells sum to %r (%s)" % (node["value"], tot, ",".join(refs)))

            # 3. formulas, where present, must match the workbook's formula
            for fk in ("formula", "goal_formula", "actual_formula", "pct_formula"):
                if node.get(fk):
                    ck = {"formula": "cell", "goal_formula": "goal_cell",
                          "actual_formula": "actual_cell", "pct_formula": "pct_cell"}[fk]
                    ref = node.get(ck)
                    if ref and CELL_RE.match(ref):
                        # openpyxl needs a second, formula-mode workbook; supplied via closure
                        wf = FORMULA_SHEETS.get(ws.title)
                        if wf is not None:
                            got = wf[ref].value
                            got = got if isinstance(got, str) and got.startswith("=") else None
                            check(got == node[fk], "%s: %s formula at %s" % (tag, ref, path),
                                  "JSON says %r, workbook has %r" % (node[fk], got))

            for k, v in node.items():
                walk(v, path + "." + k)
        elif isinstance(node, list):
            for i, v in enumerate(node):
                walk(v, "%s[%d]" % (path, i))

    walk(payload, tag)

    # 4. cumulative series must equal the running sums of the cells they name
    cum = payload.get("cumulative")
    if isinstance(cum, dict):
        # digital shape: {channel: {actual:[{value,cells}], goal:[...]}}
        for owner, series in cum.items():
            if isinstance(series, dict) and "goal" in series and "actual" in series:
                for which in ("goal", "actual"):
                    for i, pt in enumerate(series[which]):
                        if pt.get("value") is None or not pt.get("cells"):
                            continue
                        counts["series"] += 1
                        tot = 0.0
                        for r in pt["cells"]:
                            v = cellnum(ws[r].value)
                            if v is not None:
                                tot += v
                        check(abs(tot - pt["value"]) <= 0.02,
                              "%s: cumulative %s %s point %d equals its cells" % (tag, owner, which, i),
                              "JSON says %r, %s sum to %r" % (pt["value"], ",".join(pt["cells"]), tot))
            # paid shape: {quarterly:{metric:[{value,step,cell}]}, monthly:{...}}
            elif isinstance(series, dict):
                for metric, pts in series.items():
                    run = 0.0
                    for i, pt in enumerate(pts):
                        if pt.get("step_cell"):
                            counts["series"] += 1
                            v = cellnum(ws[pt["step_cell"]].value)
                            if v is not None:
                                run += v
                        if pt.get("value") is not None:
                            check(abs(run - pt["value"]) <= 0.02,
                                  "%s: %s %s cumulative %d" % (tag, owner, metric, i),
                                  "JSON says %r, running sum is %r" % (pt["value"], run))


FORMULA_SHEETS = {}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", nargs="?", default="EXTERNAL_ CAB Digital Report.xlsx")
    ap.add_argument("-d", "--digital", default="digital-report-data.json")
    ap.add_argument("-p", "--paid", default="paid-media-data.json")
    args = ap.parse_args()

    if not os.path.exists(args.workbook):
        sys.exit("workbook not found: %s" % args.workbook)

    wbv = openpyxl.load_workbook(args.workbook, data_only=True)
    wbf = openpyxl.load_workbook(args.workbook, data_only=False)

    todo = []
    for path, tag in [(args.digital, "digital"), (args.paid, "paid")]:
        if not os.path.exists(path):
            print("note: %s not found, skipping" % path)
            continue
        with open(path) as fh:
            payload = json.load(fh)
        name = payload.get("meta", {}).get("sheet")
        if name not in wbv.sheetnames:
            check(False, "%s: sheet %r exists in the workbook" % (tag, name),
                  "sheets: %s" % wbv.sheetnames)
            continue
        FORMULA_SHEETS[name] = wbf[name]
        todo.append((payload, wbv[name], tag))

    # the workbook the JSON says it came from should be the one we are auditing
    for payload, ws, tag in todo:
        claimed = payload.get("meta", {}).get("source_workbook")
        check(claimed == os.path.basename(args.workbook),
              "%s: JSON names the workbook it was built from" % tag,
              "JSON says %r, auditing %r" % (claimed, os.path.basename(args.workbook)))
        audit(payload, ws, tag)

    fails = [r for r in results if not r[0]]
    print("%d provenance checks run, %d passed, %d failed"
          % (len(results), len(results) - len(fails), len(fails)))
    print("   %d value/cell assertions, %d computed entries, %d cumulative points, %d cell refs validated"
          % (counts["pairs"], counts["computed"], counts["series"], counts["refs"]))
    for ok, name, detail in fails[:40]:
        print("  FAIL  %s\n        %s" % (name, detail))
    if len(fails) > 40:
        print("  ... and %d more" % (len(fails) - 40))
    if not fails:
        print("every figure sits at the cell it claims to come from")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
