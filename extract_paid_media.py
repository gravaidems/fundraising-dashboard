#!/usr/bin/env python3
"""
Extract the 'Paid Media Report' tab into a JSON payload where every figure
carries its source cell.

Tab layout (all rows discovered dynamically, anchored on the header rows):
  F4:G9    metric definitions key (NTL, CPA, ROAS, the three Raised figures)
  E11:N..  quarterly toplines, one row per quarter present
  E17:N..  monthly toplines for the quarter the tab is set to
  B22:N..  detail rows: Month x Channel x Placement x Objective

Note on column E of the detail block: the header reads 'Goal' but the values are
categories ("Donations SMS", "Sign Ups"), so it is carried as 'objective'.

Usage:
    python3 extract_paid_media.py [workbook.xlsx] [-o out.json]
"""

import argparse
import datetime as _dt
import json
import os
import sys

import openpyxl

SHEET = "Paid Media Report"

# key, label, column, kind
METRICS = [
    ("gross_spend", "Gross Spend", "F", "usd"),
    ("ntl", "NTL", "G", "count"),
    ("gross_cpa", "Gross CPA", "H", "usd2"),
    ("immediate_raised", "Immediate Raised", "I", "usd"),
    ("immediate_roas", "Immediate ROAS", "J", "roas"),
    ("lifetime_raised", "Lifetime Raised", "K", "usd"),
    ("lifetime_roas", "Lifetime ROAS", "L", "roas"),
    ("total_raised", "Total Raised", "M", "usd"),
    ("total_roas", "Total ROAS", "N", "roas"),
]

# Metrics that are additive across rows (safe to sum); the rest are ratios that
# must be recomputed from their components instead.
ADDITIVE = ["gross_spend", "ntl", "immediate_raised", "lifetime_raised", "total_raised"]

QUARTER_HEADER_ROW = 11
MONTH_HEADER_ROW = 17
DETAIL_HEADER_ROW = 22
DEF_ROWS = (4, 9)


def num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("$", "")
        if s in ("", "-", "#N/A", "#DIV/0!", "#VALUE!"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return float(v)


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", nargs="?", default="EXTERNAL_ CAB Digital Report.xlsx")
    ap.add_argument("-o", "--out", default="paid-media-data.json")
    args = ap.parse_args()

    if not os.path.exists(args.workbook):
        sys.exit("workbook not found: %s" % args.workbook)

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    wbf = openpyxl.load_workbook(args.workbook, data_only=False)
    if SHEET not in wb.sheetnames:
        sys.exit("sheet %r not found" % SHEET)
    ws, wsf = wb[SHEET], wbf[SHEET]

    def val(ref):
        return ws[ref].value

    def formula(ref):
        f = wsf[ref].value
        return f if isinstance(f, str) and f.startswith("=") else None

    def metrics_at(row):
        """Read the nine metric columns for one row, with cell refs."""
        out = {}
        for key, label, col, kind in METRICS:
            ref = "%s%d" % (col, row)
            out[key] = {
                "label": label, "kind": kind,
                "value": num(val(ref)), "cell": ref,
                "formula": formula(ref),
            }
        return out

    # ---- metric definitions key (F4:G9) --------------------------------
    definitions = []
    for r in range(DEF_ROWS[0], DEF_ROWS[1] + 1):
        term, text = txt(val("F%d" % r)), txt(val("G%d" % r))
        if term and text:
            definitions.append({
                "term": term.rstrip(":"), "text": text,
                "term_cell": "F%d" % r, "text_cell": "G%d" % r,
            })

    # ---- quarterly toplines --------------------------------------------
    quarters = []
    r = QUARTER_HEADER_ROW + 1
    while num(val("E%d" % r)) is not None:
        quarters.append({
            "quarter": int(num(val("E%d" % r))),
            "label": "Q%d" % int(num(val("E%d" % r))),
            "label_cell": "E%d" % r,
            "row": r,
            "metrics": metrics_at(r),
        })
        r += 1

    # ---- monthly toplines ----------------------------------------------
    months = []
    r = MONTH_HEADER_ROW + 1
    while txt(val("E%d" % r)):
        months.append({
            "label": txt(val("E%d" % r)),
            "label_cell": "E%d" % r,
            "row": r,
            "metrics": metrics_at(r),
        })
        r += 1

    # ---- detail rows ---------------------------------------------------
    detail = []
    r = DETAIL_HEADER_ROW + 1
    while txt(val("B%d" % r)):
        detail.append({
            "month": txt(val("B%d" % r)), "month_cell": "B%d" % r,
            "channel": txt(val("C%d" % r)), "channel_cell": "C%d" % r,
            "placement": txt(val("D%d" % r)), "placement_cell": "D%d" % r,
            "objective": txt(val("E%d" % r)), "objective_cell": "E%d" % r,
            "row": r,
            "metrics": metrics_at(r),
        })
        r += 1

    # ---- year-to-date rollup across the quarters present ---------------
    # Additive metrics are summed; ratios are RECOMPUTED from those sums, never
    # averaged, because averaging ratios across unequal denominators is wrong.
    ytd = {"quarters": [q["label"] for q in quarters], "metrics": {}}
    for key in ADDITIVE:
        tot, refs = 0.0, []
        for q in quarters:
            m = q["metrics"][key]
            if m["value"] is not None:
                tot += m["value"]
                refs.append(m["cell"])
        ytd["metrics"][key] = {
            "label": dict((k, l) for k, l, c, t in METRICS)[key],
            "kind": dict((k, t) for k, l, c, t in METRICS)[key],
            "value": tot, "cells": refs, "computed": True,
        }
    sp = ytd["metrics"]["gross_spend"]["value"]
    ntl = ytd["metrics"]["ntl"]["value"]
    ratio_defs = [
        ("gross_cpa", "Gross CPA", "usd2", sp, ntl, "gross_spend", "ntl"),
        ("immediate_roas", "Immediate ROAS", "roas",
         ytd["metrics"]["immediate_raised"]["value"], sp, "immediate_raised", "gross_spend"),
        ("lifetime_roas", "Lifetime ROAS", "roas",
         ytd["metrics"]["lifetime_raised"]["value"], sp, "lifetime_raised", "gross_spend"),
        ("total_roas", "Total ROAS", "roas",
         ytd["metrics"]["total_raised"]["value"], sp, "total_raised", "gross_spend"),
    ]
    for key, label, kind, numer, denom, nk, dk in ratio_defs:
        ytd["metrics"][key] = {
            "label": label, "kind": kind,
            "value": (numer / denom) if denom else None,
            "computed": True, "numerator": nk, "denominator": dk,
            "cells": ytd["metrics"][nk]["cells"] + ytd["metrics"][dk]["cells"],
        }

    # ---- cumulative series --------------------------------------------
    # Running sums of the additive metrics, at both quarterly and monthly grain.
    def cumulate(rows, keys):
        out = {}
        for key in keys:
            run, series = 0.0, []
            for row in rows:
                v = row["metrics"][key]["value"]
                if v is not None:
                    run += v
                series.append({
                    "label": row["label"], "value": run,
                    "step": v, "cell": row["metrics"][key]["cell"],
                })
            out[key] = series
        return out

    ACCUM_KEYS = ["gross_spend", "ntl", "total_raised"]
    cumulative = {
        "quarterly": cumulate(quarters, ACCUM_KEYS),
        "monthly": cumulate(months, ACCUM_KEYS),
    }

    payload = {
        "meta": {
            "title": txt(val("B2")),
            "sheet": SHEET,
            "source_workbook": os.path.basename(args.workbook),
            "extracted_at": _dt.datetime.now().isoformat(timespec="seconds"),
            "year": {"value": num(val("C4")), "cell": "C4"},
            "quarter": {"value": num(val("C5")), "cell": "C5"},
            "client": {"value": txt(val("C7")), "cell": "C7"},
            "lead": {"value": txt(val("C8")), "cell": "C8"},
        },
        "metric_order": [m[0] for m in METRICS],
        "metric_labels": dict((k, l) for k, l, c, t in METRICS),
        "metric_kinds": dict((k, t) for k, l, c, t in METRICS),
        "metric_columns": dict((k, c) for k, l, c, t in METRICS),
        "additive": ADDITIVE,
        "accum_keys": ACCUM_KEYS,
        "definitions": definitions,
        "quarters": quarters,
        "months": months,
        "detail": detail,
        "ytd": ytd,
        "cumulative": cumulative,
        "channels": sorted(set(d["channel"] for d in detail if d["channel"])),
        "objectives": sorted(set(d["objective"] for d in detail if d["objective"])),
    }

    with open(args.out, "w") as fh:
        json.dump(payload, fh, indent=2)
    print("wrote %s (%d quarters, %d months, %d detail rows, %d channels)"
          % (args.out, len(quarters), len(months), len(detail), len(payload["channels"])))


if __name__ == "__main__":
    main()
