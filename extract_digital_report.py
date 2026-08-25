#!/usr/bin/env python3
"""
Extract the 'Digital Report' tab from the CAB Digital Report workbook into a
JSON payload where every figure carries the source cell that produced it.

The workbook's own values originate from a Google Sheet via IMPORTRANGE and are
cached as static values in the .xlsx, so we read with data_only=True and also
inspect the formula layer (data_only=False) to record derivations.

Usage:
    python3 extract_digital_report.py [workbook.xlsx] [-o out.json]
"""

import argparse
import datetime as _dt
import json
import os
import sys

import openpyxl

SHEET = "Digital Report"

# Row layout of the per-channel grid on the Digital Report tab.
CHANNEL_ROWS = [
    ("Email", 14),
    ("Recurring", 15),
    ("Ads", 16),
    ("Website", 17),
    ("Tandem", 18),
    ("SMS Broadcast", 19),
    ("SMS P2P", 20),
    ("Social", 21),
]

# Column layout: each period occupies a (goal, actual, pct) column triple.
PERIODS = [
    ("quarter", "Quarter", "C", "D", "E"),
    ("july", "July", "G", "H", "I"),
    ("august", "August", "J", "K", "L"),
    ("september", "September", "M", "N", "O"),
]

MONTH_KEYS = ["july", "august", "september"]

# Hidden helper columns Q:S feed the quarterly goal via C = SUM(Q:S).
HELPER_COLS = {"july": "Q", "august": "R", "september": "S"}

OVERALL_ROWS = [
    ("total_raised", "Total Raised", 23),
    ("total_donations", "Total Donations", 24),
    ("average_donation", "Average Donation", 25),
    ("total_finance", "Total Finance", 26),
]


def num(v):
    """Coerce a cell value to float, or None for blanks and the '-' sentinel."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "-", "#N/A", "#DIV/0!"):
            return None
        try:
            return float(s.replace(",", "").replace("$", "").replace("%", ""))
        except ValueError:
            return None
    if isinstance(v, bool):
        return None
    return float(v)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "workbook",
        nargs="?",
        default="EXTERNAL_ CAB Digital Report.xlsx",
        help="path to the .xlsx workbook",
    )
    ap.add_argument("-o", "--out", default="digital-report-data.json")
    args = ap.parse_args()

    if not os.path.exists(args.workbook):
        sys.exit("workbook not found: %s" % args.workbook)

    wb_v = openpyxl.load_workbook(args.workbook, data_only=True)
    wb_f = openpyxl.load_workbook(args.workbook, data_only=False)
    if SHEET not in wb_v.sheetnames:
        sys.exit("sheet %r not found; sheets are %s" % (SHEET, wb_v.sheetnames))
    ws = wb_v[SHEET]
    wsf = wb_f[SHEET]

    def val(ref):
        return ws[ref].value

    def formula(ref):
        f = wsf[ref].value
        return f if isinstance(f, str) and f.startswith("=") else None

    # ---- header / metadata -------------------------------------------------
    meta = {
        "title": val("B4"),
        "client": {"value": val("C6"), "cell": "C6"},
        "email_lead": {"value": val("C7"), "cell": "C7"},
        "year": {"value": num(val("C9")), "cell": "C9"},
        "quarter": {"value": num(val("C10")), "cell": "C10"},
        "dynamic_goals": {
            "july": {"value": bool(val("O9")), "cell": "O9"},
            "august": {"value": bool(val("O10")), "cell": "O10"},
            "label": val("N8"),
        },
        "source_workbook": os.path.basename(args.workbook),
        "extracted_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "sheet": SHEET,
    }

    # ---- per-channel grid --------------------------------------------------
    channels = []
    for name, row in CHANNEL_ROWS:
        entry = {
            "name": ws["B%d" % row].value or name,
            "name_cell": "B%d" % row,
            "row": row,
            "periods": {},
            "helpers": {},
        }
        for key, label, gc, ac, pc in PERIODS:
            g_ref, a_ref, p_ref = "%s%d" % (gc, row), "%s%d" % (ac, row), "%s%d" % (pc, row)
            entry["periods"][key] = {
                "label": label,
                "goal": num(val(g_ref)),
                "goal_cell": g_ref,
                "actual": num(val(a_ref)),
                "actual_cell": a_ref,
                "pct": num(val(p_ref)),
                "pct_cell": p_ref,
                "pct_raw": val(p_ref),
                "goal_formula": formula(g_ref),
                "actual_formula": formula(a_ref),
                "pct_formula": formula(p_ref),
            }
        for mkey, col in HELPER_COLS.items():
            ref = "%s%d" % (col, row)
            entry["helpers"][mkey] = {
                "value": num(val(ref)),
                "cell": ref,
                "formula": formula(ref),
            }
        # A channel with no goal and no actual anywhere is dormant this quarter.
        q = entry["periods"]["quarter"]
        entry["dormant"] = not (q["goal"] or q["actual"])
        channels.append(entry)

    # ---- overall block -----------------------------------------------------
    overall = {}
    for key, label, row in OVERALL_ROWS:
        item = {"label": ws["B%d" % row].value or label, "label_cell": "B%d" % row,
                "row": row, "periods": {}}
        for pkey, plabel, gc, ac, pc in PERIODS:
            g_ref, a_ref, p_ref = "%s%d" % (gc, row), "%s%d" % (ac, row), "%s%d" % (pc, row)
            item["periods"][pkey] = {
                "label": plabel,
                "goal": num(val(g_ref)),
                "goal_cell": g_ref,
                "actual": num(val(a_ref)),
                "actual_cell": a_ref,
                "pct": num(val(p_ref)),
                "pct_cell": p_ref,
            }
        overall[key] = item

    # ---- cumulative series (computed, not read) ----------------------------
    # Cumulative actual stops at the last month with reported actuals; the
    # cumulative goal runs the full quarter so the remaining gap stays visible.
    cumulative = {}
    for ch in channels:
        ca, cg, run_a, run_g = [], [], 0.0, 0.0
        for mkey in MONTH_KEYS:
            p = ch["periods"][mkey]
            g = p["goal"] or 0.0
            run_g += g
            cg.append({"month": p["label"], "value": run_g,
                       "cells": [ch["periods"][m]["goal_cell"] for m in MONTH_KEYS[:len(cg) + 1]]})
            if p["actual"] is None:
                ca.append({"month": p["label"], "value": None, "pending": True})
            else:
                run_a += p["actual"]
                ca.append({"month": p["label"], "value": run_a, "pending": False,
                           "cells": [ch["periods"][m]["actual_cell"] for m in MONTH_KEYS[:len(ca) + 1]]})
        cumulative[ch["name"]] = {"actual": ca, "goal": cg}

    payload = {
        "meta": meta,
        "channels": channels,
        "overall": overall,
        "cumulative": cumulative,
        "month_keys": MONTH_KEYS,
    }

    with open(args.out, "w") as fh:
        json.dump(payload, fh, indent=2)
    print("wrote %s (%d channels)" % (args.out, len(channels)))
    return payload


if __name__ == "__main__":
    main()
